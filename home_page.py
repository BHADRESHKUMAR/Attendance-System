from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook
import openpyxl
from openpyxl import Workbook
import xlrd
import datetime
import face_ditection
import pandas as pd
root = Tk()
f = Frame(root)
frame1 = Frame(root)
frame2 = Frame(root)
frame3 = Frame(root)
root.title("Attendance")
root.geometry("830x395")
root.configure(background="Light blue")

scrollbar = Scrollbar(root)
scrollbar.pack(side=RIGHT, fill=Y)

sname = StringVar()
fname = StringVar()
id = StringVar()
lecture = StringVar()
sheet_data = []
row_data = []


def emp_dict(*args):  # To add a new entry and check if entry already exist in excel sheet
    # print("done")
    workbook_name = "home.xlsx"
    workbook = xlrd.open_workbook(workbook_name)
    worksheet = workbook.sheet_by_index(0)

    wb = load_workbook(workbook_name)
    page = wb.active

    p = 0
    for i in range(worksheet.nrows):
        for j in range(worksheet.ncols):
            cellvalue = worksheet.cell_value(i, j)
            # print(cellvalue)
            sheet_data.append([])
            sheet_data[p] = cellvalue
            p += 1
    # print(sheet_data)
    fl = sname.get()
    fsl = fl.lower()
    ll = fname.get()
    lsl = ll.lower()
    if (fsl and lsl) in sheet_data:
        print("found")
        messagebox.showerror("Error", "Attendance is already taken")
        wb.save(filename=workbook_name)

    else:
        print("not found")
        for info in args:
            page.append(info)
        messagebox.showinfo("Done", "Now you can able to take the attendance")
        wb.save(filename=workbook_name)
        face_ditection.face()
        date = datetime.datetime.now()
        a = date.date()
        print(a)
        file1 = lsl+".xlsx"
        wb = openpyxl.load_workbook(file1)
        sheet = wb['Sheet1']
        # b = sheet[('B'+1)+'2'].value.date()
        data2 = 1

        # Workbook is created

        # add_sheet is used to create sheet.
        for i in range(2, 20):
            if (a == sheet.cell(row=9, column=i).value.date()):
                data2 = i
                break
            else:
                pass

        if (face_ditection.bhadresh != 0):
            sheet.cell(row=10, column=i).value = 1
        else:
            sheet.cell(row=10, column=i).value = 0
        if (face_ditection.parth != 0):
            sheet.cell(row=11, column=i).value = 1
        else:
            sheet.cell(row=11, column=i).value = 0
        if (face_ditection.niraj != 0):
            sheet.cell(row=13, column=i).value = 1
        else:
            sheet.cell(row=13, column=i).value = 0
        if (face_ditection.chintan != 0):
            sheet.cell(row=12, column=i).value = 1
        else:
            sheet.cell(row=12, column=i).value = 0
        wb.save(file1)

def submit():  # to append all data and add entries on click the button
    a = " "
    f = sname.get()
    f1 = f.lower()
    l = fname.get()
    l1 = l.lower()
    d = lecture.get()
    d1 = d.lower()
    list1 = list(a)
    list1.append(f1)
    list1.append(l1)
    list1.append(d1)
    emp_dict(list1)


def add_info():  # for taking user input to add the enteries
    frame2.pack_forget()
    frame3.pack_forget()
    subject_name = Label(frame1, text="Subject Name : ", bg="Light blue", fg="purple")
    subject_name.grid(row=1, column=1, padx=20)
    sname.set("Select Option")
    e1 = OptionMenu(frame1, sname,"Operating ","Digital Electronics","Maths","Software Group Project",command=func)
    e1.grid(row=1, column=2, padx=20)
    e1.focus()

    emp_lecture = Label(frame1, text="Lecture : ", bg="Light gray", fg="black")
    emp_lecture.grid(row=3, column=1, padx=20)
    lecture.set("Select Option")
    e4 = OptionMenu(frame1, lecture, "Theory", "Practical")
    e4.grid(row=3, column=2, padx=20)

    button4 = Button(frame1, text="Submit", width = 8 ,command=submit)
    button4.grid(row=5, column=1, pady=20)

    frame1.configure(background="Light gray")


    button6 = Button(frame1, text="Close",width = 8, command=root.destroy)
    button6.grid(row=5, column=2, pady=20)
    frame1.configure(background="Light gray")
    frame1.pack(pady=20)

def func(value):
    s1 = value
    faculty = Label(frame1, text="Faculty Name : ", bg="Light gray", fg="black")
    faculty.grid(row=2, column=1, padx=20)
    fname.set("Select Name")
    if (s1 == "JAVA Programming"):
        e2 = OptionMenu(frame1, fname, "Minal_Maniar", "Mrugendrasinh_Rahevar")
    elif (s1 == "Digital Electronics"):
        e2 = OptionMenu(frame1, fname,  "Ronak_Patel", "Aniruddh_Fataniya")
    elif (s1 == "Maths"):
        e2 = OptionMenu(frame1, fname, "Niru_Patel", "Yogeshwari_Patel")
    else:
        e2 = OptionMenu(frame1, fname, "Vashali_Mewada", "Dhruti_Pandya")
    e2.grid(row=2, column=2, padx=20)


def clear_all():  # for clearing the entry widgets
    frame1.pack_forget()
    frame2.pack_forget()
    frame3.pack_forget()

# Main window buttons and labels

label1 = Label(root, text="Attendence")
label1.config(font=('Italic', 40, 'bold'), justify=CENTER, background="Black", fg="white", anchor="center")
label1.pack(fill=X)
add_info()


root.mainloop()
